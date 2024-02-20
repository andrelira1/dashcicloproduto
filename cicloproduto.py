import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from consulta import *
from streamlit_option_menu import option_menu
from PIL import Image
from openpyxl import Workbook
from data_handler import *
import io
from tkinter import filedialog

dfc_rf, df_rc, ultima_atualizacao = atualiza()

###--- Esconde alguns Styles, como o rodapé.
hide_footer = """
          <style>
          #MainMenu {
              visibility: visible;
          }
          footer {
            visibility: hidden;
          }
          header {
            visibility: hidden;
          }
          footer:after{
              content:'Powered by André Lira';
              visibility: visible;
              display:block;
              position:relative;
              color:gray;
          }
          </style>
          """

icon = Image.open('images/icon.png')
st.set_page_config(page_title= "Ciclo Produto",
                   page_icon=icon,
                   layout="wide")

###-----------------TÍTULO-----------------###
sTitulo = """
<style>
.custom-title {
    background-color: #23998E;
    padding: 10px;
    border: 2px solid #d0d0d0;
    border-radius: 15px;
    font-size: 32px;
    text-align: center;
    color: #d0d0d0; //#333333;
    font-weight: bold;
}
</style>
"""
# Renderizar o CSS personalizado
st.markdown(sTitulo, unsafe_allow_html=True)
# Título da página com estilo personalizado
st.markdown('<div class="custom-title">Painel Produtor</div>', unsafe_allow_html=True)

###-----------------LAYOUT-----------------###
with open("style.css") as f:
    st.markdown(f'<style>{f.read()}</style>', unsafe_allow_html=True)

###----------------SIDEBAR---------------###
img_sidebar = "static/logo2.png"
st.sidebar.image(img_sidebar, width=10, use_column_width=True)


###---------------FILTROS---------------###
with st.sidebar:
    menu_selected = option_menu(None, ['Início', 'Produtor', 'Analítico', 'Comercial'],
                                icons=['house', 'search', 'file-earmark', 'basket2-fill'],
                                default_index=0,
                                styles={"container": {"background-color": "#23998E", "padding": "0!important", "border-radius": "5px"},
                                        "nav-link": {"color": "white", "font-family": "'Open Sans', sans-serif"},
                                        "nav-link-selected": {"color": "white", "background-color": "#1D5E69", "font-family": "'Open Sans', sans-serif"},
                                        })
                                        #COLORES ==> Color #061638 -> #1D5E69 e Background white ->
with st.sidebar.expander("Filtros"):
 ##--Filtro da Safra
    fSafra = sorted((dfc_rf['SAFRA'].value_counts().index).tolist())
    fSafra.insert(0, 'TODOS')
    saf_selecao = st.selectbox("Safra: ", fSafra)
    if saf_selecao != 'TODOS':
        dfc_rf = dfc_rf.query("SAFRA == @saf_selecao")

 ##--Filtro do Controle
    controle = sorted((dfc_rf['CONTROLE'].value_counts().index).tolist())
    controle.insert(0, 'TODOS')
    ctr_selecao = st.selectbox("Controle: ", controle)
    if ctr_selecao != 'TODOS':
        dfc_rf = dfc_rf.query("CONTROLE == @ctr_selecao")

 ##--Filtro do NF VENDA
    nfvenda = sorted((dfc_rf['NF_VENDA'].value_counts().index).tolist())
    nfvenda.insert(0, 'TODOS')
    nfv_selecao = st.selectbox("NF Venda: ", nfvenda)
    if nfv_selecao != 'TODOS':
        dfc_rf = dfc_rf.query("NF_VENDA == @nfv_selecao")

##--Filtro do Cliente
    fCliente = sorted((dfc_rf['CLIENTE'].value_counts().index).tolist())
    fCliente.insert(0, 'TODOS')
    cli_selecao = st.selectbox("Cliente: ", fCliente)
    if cli_selecao != 'TODOS':
        dfc_rf = dfc_rf.query("CLIENTE == @cli_selecao")

##--Filtro do NF COMPRA
    nfcompra = sorted((dfc_rf['NF_COMPRA'].value_counts().index).tolist())
    nfcompra.insert(0, 'TODOS')
    nfc_selecao = st.selectbox("NF Compra: ", nfcompra)
    if nfc_selecao != 'TODOS':
        dfc_rf = dfc_rf.query("NF_COMPRA == @nfc_selecao")
  
##--Filtro do Produtor
    opcaoFA = sorted((dfc_rf['PRODUTOR'].value_counts().index).tolist())
    opcaoFA.insert(0, 'TODOS')
    fa_selecao = st.selectbox("Produtor: ", opcaoFA)
    if fa_selecao != 'TODOS':
        dfc_rf = dfc_rf.query("PRODUTOR == @fa_selecao")


    if st.button('Atualizar'):
        dfc_rf, df_rc, ultima_atualizacao = atualiza()
    #   st.experimental_rerun()
    st.markdown(f'##### Última atualização: {ultima_atualizacao}')

###----------------VARIÁVEIS---------------###

##--Total Pallets
    TotalPallet = dfc_rf['PALLET'].count()
##--Volume Por Safra
    VolumeSafra = dfc_rf.groupby('SAFRA')['QTD_VENDA_KG'].sum()
##--Pallets por NF Venda
    PalletNFV = dfc_rf.groupby('NF_VENDA')['PALLET'].count()
##--Volume Por Fundo Agrícola
    VolumeFA = dfc_rf.groupby('PRODUTOR')['QTD_VENDA_KG'].sum()
##--Volume Por NF Venda
    VolumeNFVKG = dfc_rf.groupby('NF_VENDA')['QTD_VENDA_KG'].max()
##--Volume Por NF Venda
    VolumeNFVCX = dfc_rf.groupby('NF_VENDA')['QTD_VENDA_CX'].max()
##--Volume Venda por Controle
    VolumeVCtrl = dfc_rf.groupby('CONTROLE')['QTD_VENDA_KG'].sum()
##--Volume Compra por Controle
    VolumeCCtrl = dfc_rf.groupby('CONTROLE')['NFC_QTD_ITEM'].sum()
##--Dados para exportar para excel
    dadosexcel = dfc_rf.copy()

###----------------FIGURAS-----------------###


    dfrf1_1 = dfc_rf.groupby(['SAFRA', 'CONTROLE'])['PALLET'].count().reset_index()

    graf1 = go.Figure()
    graf1.add_trace(go.Bar(x=dfc_rf['MODALIDADE'],
                        #y=dfc_rf['PALLET'],
                        y=dfrf1_1['PALLET'],
                        text=dfrf1_1['PALLET'],
                        textposition='auto',
                        orientation='v'))
    graf1.update_layout(xaxis={'title': None, 'fixedrange': True},
                        yaxis={'title': None, 'fixedrange': True},
                        font_family='Open Sans',
                        hovermode=False,
                        xaxis_tickangle=-45)

##-- Graf 2 - Volume por Safra
    dfrf2 = dfc_rf['QTD_VENDA_KG']
    dfrf2 = dfc_rf['CONTROLE']

    graf2 = go.Figure()
    graf2.add_trace(go.Pie(labels=dfrf2,
                        values=dfrf2,
                        hole=.6,
                        hovertemplate=None))
    graf2.update_layout(uniformtext_minsize=12,
                        uniformtext_mode='hide',
                        font_family='Open Sans',
                        legend_title_text="CONTROLE",
                        legend_itemclick=False,
                        legend_itemdoubleclick=False)

    
    ##st.write(dfc2.tail(50))
    dfrf3 = dfc_rf.groupby(['SAFRA', 'CONTROLE'])['QTDCXPALLET'].count().reset_index()
    # Cria um gráfico de linha com Plotly
    
    graf3 = px.line(dfrf3, x="CONTROLE", y="QTDCXPALLET") #, color="PALLET")

##-- Graf 4 - Volume por Safra
    dfrf4 = dfc_rf['CONTROLE'].value_counts().iloc[::-1]

    graf4 = go.Figure()
    graf4.add_trace(go.Bar(x=dfrf4,
                        y=dfrf4.index,
                        orientation='v',
                        text=dfc_rf))
    graf4.update_layout(xaxis={'title': None, 'fixedrange': True},
                        yaxis={'title': None, 'fixedrange': True},
                        font_family='Open Sans',
                        hovermode=False)

##-- Graf 5 - Volume por Safra
    dfrf5 = dfc_rf['PALLET'].value_counts().iloc[::-1]

    graf5 = go.Figure()
    graf5.add_trace(go.Bar(x=dfrf5,
                        y=dfrf5.index,
                        orientation='v',
                        text=dfrf5))
    graf5.update_layout(xaxis={'title': None, 'fixedrange': True},
                        yaxis={'title': None, 'fixedrange': True},
                        font_family='Open Sans',
                        hovermode=False)

##-- Graf 6 - Volume por Safra
    dfrf6 = dfc_rf['QTD_VENDA_KG'].value_counts()

    graf6 = go.Figure()
    graf6.add_bar(x=dfc_rf['NF_COMPRA'].value_counts().index,
                y=dfc_rf['CONTROLE'],
                marker_color=['#0068c9', '#83c9ff'],
                hovertemplate='<br><b>Total Vendido:</b> %{y}<extra></extra>')

    graf6.update_layout(xaxis_tickangle=0,
                    xaxis={'title': None, 'fixedrange': True},
                    yaxis={'title': None, 'fixedrange': True},
                    font_family='Open Sans',
                    legend_title_text="GRUPO",
                    legend_itemclick=False,
                    legend_itemdoubleclick=False)

##-- Graf 7 - Qtd Pallet por Safra e por Controle
    dfrf7 = dfc_rf.groupby(['SAFRA', 'CONTROLE'])['PALLET'].count().reset_index()
    
    graf7 = go.Figure()
    graf7.add_trace(go.Bar(x=dfrf7['SAFRA'] + ' - ' + dfrf7['CONTROLE'].astype(str),
                           y=dfrf7['PALLET'],
                           text=dfrf7['PALLET'],
                           textposition='auto'))
    graf7.update_layout(xaxis={'title': 'SAFRA - CONTROLE', 'fixedrange': True},
                    yaxis={'title': 'Quantidade de Pallets', 'fixedrange': True},
                    font_family='Open Sans',
                    hovermode=False,
                    title='Quantidade de Pallets por SAFRA e CONTROLE',
                    xaxis_tickangle=-45)


###---------------------COLUNAS----------------------###

###----- ÁREA INÍCIO -----###
if menu_selected == 'Início':

##-- Visual dos Filtros
  col1, col2, col3 = st.columns([20,20,20])
  with col1:
    st.write("NF Venda: ", nfv_selecao)
    st.write("Cliente: ", cli_selecao)

  with col2:
    st.write("Safra: ", saf_selecao)
    st.write("Controle: ", ctr_selecao)

  with col3:
    st.write("NF Compra: ", nfc_selecao)
    st.write("Produtor: ", fa_selecao)

##-- Cards Gerais Linha 1
  col1, col2, col3 = st.columns([20,20,20])
  with col1:
    st.metric(label="Volume NF Venda (Kg)", value=VolumeNFVKG.iloc[0], delta=123, delta_color='off',)

  with col2:
    st.metric(label="Total Pallets", value=TotalPallet, delta=123, delta_color='off',)
    
  with col3:
    st.metric(label="Volume Fornecedor (Kg)", value=VolumeFA.iloc[0], delta=123, delta_color='off',)

##-- Cards Gerais Linha 2
  col1, col2, col3 = st.columns([20,20,20])
  with col1:
    st.metric(label="Volume NF Venda (Cx)", value=VolumeNFVCX.iloc[0], delta=123, delta_color='off',)

  with col2:
    st.metric(label="Volume Venda Por Controle", value=VolumeVCtrl.iloc[0], delta=123, delta_color='off',)
    
  with col3:
    st.metric(label="Volume Compra Por Controle", value=VolumeCCtrl.iloc[0], delta=123, delta_color='off',)

###----- ÁREA PRODUTOR -----###
elif menu_selected == 'Produtor':

##--Gráficos linha 1
  col1, col2, col3 = st.columns([20,20,20])
  with col1:

    st.markdown("#### G1 - QTD PALLET por MODALIDADE")
    st.plotly_chart(graf1, use_container_width=True,
                    theme='streamlit', config={'diplayModeBar': False})

  with col2:

    st.markdown("#### G2 - QTD CAIXA por CONTROLE")

    st.plotly_chart(graf2, use_container_width=True,
                    theme='streamlit', config={'diplayModeBar': False})

  with col3:
    st.markdown("#### Gráfico 7")
    st.plotly_chart(graf7, use_container_width=True,
                   theme='streamlit', config={'diplayModeBar': False})
    
##--Gráficos linha 2
  col1, col2 = st.columns([3,2])
  with col1:
    st.markdown("#### Gráfico 4")
    st.plotly_chart(graf4, use_container_width=True,
                    theme='streamlit', config={'diplayModeBar': False})
  with col2:
    st.markdown("#### Gráfico 5")
    st.plotly_chart(graf5, use_container_width=True,
                    theme='streamlit', config={'diplayModeBar': False})
    
##--Gráficos linha 3
  col1, col2 = st.columns([3,2])
  with col1:
    st.markdown("#### G3 - QTD CAIXA por CONTROLE")
    st.plotly_chart(graf3, use_container_width=True,
                    theme='streamlit', config={'diplayModeBar': False})
  with col2:
    st.markdown("#### Gráfico 6")
    st.plotly_chart(graf6, use_container_width=True,
                    theme='streamlit', config={'diplayModeBar': False})


  col1, col2, col3 = st.columns([20,20,20])
  with col1:
      st.markdown("#### Coluna Produtor 1")
  with col2:
      st.markdown("#### Coluna Produtor 2")
  with col3:
      st.markdown("#### Coluna Produtor 3")

###----- ÁREA NF VENDA -----###
elif menu_selected == 'Analítico':
###-----------------TÍTULO-----------------###
  sTituloRel = """
  <style>
  .custom-titler {
      background-color: #23998E;
      padding: 6px;
      border: 2px solid #d0d0d0;
      border-radius: 15px;
      font-size: 28px;
      text-align: center;
      color: #d0d0d0; //#333333;
      font-weight: bold;
      margin-bottom: 20px;
  }
  </style>
  """
  # Renderizar o CSS personalizado
  st.markdown(sTituloRel, unsafe_allow_html=True)
  # Título da página com estilo personalizado
  st.markdown('<div class="custom-titler">Relatório Analítico</div>', unsafe_allow_html=True)


  col1, col2 = st.columns([1000,1])
  with col1:
   st.write(dadosexcel)

  ##--Criando um novo arquivo Excel
  def export_to_excel(dadosexcel):
     try:
         workbook = Workbook()
         sheet = workbook.active
         
         ##-- Escrevendo os cabeçalhos das colunas no Excel (primeira linha)
         for coluna_idx, coluna_nome in enumerate(dadosexcel.columns, start=1):
            sheet.cell(row=1, column=coluna_idx, value=coluna_nome)

         ##-- Escrevendo os dados no Excel
         for linha_idx, row in enumerate(dadosexcel.iterrows(), start=1):
             for coluna_idx, value in enumerate(row[1], start=1):
               sheet.cell(row=linha_idx +1, column=coluna_idx, value=value)
         ##-- Criando um objeto BytesIO para guardar o arquivo Excel em formato binário
         excel_file = io.BytesIO()
         workbook.save(excel_file)
         ##-- Reinicia a posição do objeto BytesIO para o início do arquivo
         excel_file.seek(0)

        ##-- Abrindo o diálogo de salvamento (Save As) e permitindo que o usuário escolha o local
         localarquivo = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])

         if localarquivo:
         ##-- Salvando o arquivo e criando o link para download
            with open(localarquivo, "wb") as f:
                #workbook.save(f)
                f.write(excel_file.getvalue()) # Grava o conteúdo do BytesIO no arquivo
            
            st.success("Arquivo salvo com sucesso em: {}".format(localarquivo))
         
     except Exception as e:
         st.error("Ocorreu um erro ao gerar o arquivo Excel. Detalhes: {}".format(str(e)))

  ##-- Botão para extrair os dados para Excel
  if st.button('Exportar para Excel', on_click=export_to_excel, args=(dadosexcel,)):
    pass 

  with col2:
   st.markdown("")

elif menu_selected == 'Comercial':
   ###-----------------TÍTULO-----------------###
  sTituloRel = """
  <style>
  .custom-titler {
      background-color: #23998E;
      padding: 6px;
      border: 2px solid #d0d0d0;
      border-radius: 15px;
      font-size: 28px;
      text-align: center;
      color: #d0d0d0; //#333333;
      font-weight: bold;
      margin-bottom: 20px;
  }
  </style>
  """
  # Renderizar o CSS personalizado
  st.markdown(sTituloRel, unsafe_allow_html=True)
  # Título da página com estilo personalizado
  st.markdown('<div class="custom-titler">Resumo Comercial</div>', unsafe_allow_html=True)
  excel_rc = df_rc
  col1, col2 = st.columns([1000,1])
  with col1:
   #  st.write(df_rc)

   st.write(excel_rc)

  ##--Criando um novo arquivo Excel
  def export_to_excel(excel_rc):
     try:
         workbook = Workbook()
         sheet = workbook.active
         
         ##-- Escrevendo os cabeçalhos das colunas no Excel (primeira linha)
         for coluna_idx, coluna_nome in enumerate(excel_rc.columns, start=1):
            sheet.cell(row=1, column=coluna_idx, value=coluna_nome)

         ##-- Escrevendo os dados no Excel
         for linha_idx, row in enumerate(excel_rc.iterrows(), start=1):
             for coluna_idx, value in enumerate(row[1], start=1):
               sheet.cell(row=linha_idx +1, column=coluna_idx, value=value)
         ##-- Criando um objeto BytesIO para guardar o arquivo Excel em formato binário
         excel_file_rc = io.BytesIO()
         workbook.save(excel_file_rc)
         ##-- Reinicia a posição do objeto BytesIO para o início do arquivo
         excel_file_rc.seek(0)

        ##-- Abrindo o diálogo de salvamento (Save As) e permitindo que o usuário escolha o local
         localarquivo_rc = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])

         if localarquivo_rc:
         ##-- Salvando o arquivo e criando o link para download
            with open(localarquivo_rc, "wb") as f:
                #workbook.save(f)
                f.write(excel_file_rc.getvalue()) # Grava o conteúdo do BytesIO no arquivo
            
            st.success("Arquivo salvo com sucesso em: {}".format(localarquivo_rc))

     except Exception as e:
         st.error("Ocorreu um erro ao gerar o arquivo Excel. Detalhes: {}".format(str(e)))

  ##-- Botão para extrair os dados para Excel
  if st.button('Exportar para Excel', on_click=export_to_excel, args=(excel_rc,)):
    pass #st.success("Arquivo gerado com sucesso!")


  with col2:
   st.markdown("")

st.markdown(hide_footer, unsafe_allow_html=True)